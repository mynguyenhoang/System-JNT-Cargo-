# -*- coding: utf-8 -*-
"""
HE THONG DU LIEU VAN HANH QC - BAN CLOUD (SUPABASE)
"""
import datetime
import io
import pandas as pd
import psycopg2
import requests
from openpyxl.utils import get_column_letter
import streamlit as st

SO_DONG_HIEN_THI = 500

st.set_page_config(
    page_title="QC Operations",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 1rem; padding-bottom: 2rem; }
[data-testid="metric-container"] {
    background: #ffffff;
    border: 1px solid #dee2e6;
    border-radius: 6px;
    padding: 1rem 1.2rem !important;
}
.stFormSubmitButton button[kind="primaryFormSubmit"],
.stButton button[kind="primary"] {
    background-color: #1a3a6b !important;
    border: none !important;
    border-radius: 5px !important;
    font-weight: 600 !important;
    letter-spacing: 0.03em !important;
}
.stDownloadButton button {
    background-color: #1a3a6b !important;
    color: white !important;
    border: none !important;
    border-radius: 5px !important;
    font-weight: 600 !important;
}
[data-testid="stDataFrame"] { border: 1px solid #dee2e6; border-radius: 6px; overflow: hidden; }
</style>
""", unsafe_allow_html=True)

st.markdown(
    """
    <div style="
        background:#1a3a6b;
        padding:14px 24px;
        border-radius:6px;
        margin-bottom:20px;
        display:flex;
        align-items:baseline;
        gap:14px;
    ">
        <span style="color:#fff;font-size:17px;font-weight:700;letter-spacing:.04em;">
            HỆ THỐNG DỮ LIỆU VẬN HÀNH QC
        </span>
        <span style="color:#93b4da;font-size:13px;">Hub Operations Dashboard</span>
    </div>
    """,
    unsafe_allow_html=True,
)

DB_URL = st.secrets.get("DB_URL", "")
if not DB_URL:
    st.error("Chưa khai báo DB_URL trong Secrets của app.")
    st.stop()

FEISHU_APP_ID             = st.secrets.get("FEISHU_APP_ID", "cli_a9456e412bb89bce")
FEISHU_APP_SECRET         = st.secrets.get("FEISHU_APP_SECRET", "BwSAuHHsv2woEdIGTqJoKboH6i1i7qBB")

# Sheet COT dùng trực tiếp để tính Ontime xếp hàng.
# Có thể ghi đè bằng Secrets; mặc định dùng đúng sheet COT của bộ cào JFS.
FEISHU_COT_SPREADSHEET_TOKEN = st.secrets.get(
    "FEISHU_COT_SPREADSHEET_TOKEN",
    "PhCAsHyCXh7BEStPRSwcxHMCnKe"
)
FEISHU_COT_SHEET_ID = st.secrets.get(
    "FEISHU_COT_SHEET_ID",
    "nJppxx"
)

# Token này chỉ dùng cho các chức năng Feishu cũ khác (nếu có).
FEISHU_SPREADSHEET_TOKEN  = st.secrets.get(
    "FEISHU_SPREADSHEET_TOKEN",
    "LXeHseOdthPKm0tnpChcjonKnkf"
)


def ket_noi():
    return psycopg2.connect(DB_URL)


BANG = {
    "Quét hàng (lên / xuống xe)": ("quet_hang", "Ngày vận hành"),
    "Tiến độ xếp hàng":             ("xep_hang",  "Ngày vận hành"),
}

TEN_BO_PHAN_XEP = {"SHDC": "SH DC", "HCM": "HCM HUB", "BN": "BN HUB"}

# Tên Hub trên giao diện -> tên Hub thực tế trong bảng bao_cao Supabase.
HUB_BAO_CAO = {
    "HCM": "HCM HUB",
    "BN": "BN HUB",
    "SH DC": "SH DC",
}

LOAI_TUYEN_EN = {"Tuyến chính": "Linehaul", "Tuyến phụ": "Shuttle"}
LOAI_TUYEN_NHAN = {
    "Tuyến chính": "Tuyến chính (Linehaul)",
    "Tuyến phụ":   "Tuyến phụ (Shuttle)",
}
NHAN_TOI_LOAI_TUYEN = {v: k for k, v in LOAI_TUYEN_NHAN.items()}


def _ngay(s):
    try:
        d = pd.to_datetime(s)
        if pd.isna(d):
            return None
        return d.date()
    except Exception:
        return None


def _chuan_hoa_ma(series):
    s = series.astype(str)
    s = s.str.strip()
    s = s.str.replace(r"\.0$", "", regex=True)
    s = s.str.replace(r"\s+", "", regex=True)
    s = s.str.upper()
    s = s.replace({"NAN": "", "NONE": ""})
    return s


@st.cache_data(ttl=300)
def lay_hub():
    con = ket_noi()
    df = pd.read_sql('SELECT DISTINCT "Hub" FROM quet_hang ORDER BY 1', con)
    con.close()
    return df["Hub"].tolist()


@st.cache_data(ttl=300)
def khoang_ngay(table, date_col):
    con = ket_noi()
    df = pd.read_sql(
        f'SELECT MIN("{date_col}") AS a, MAX("{date_col}") AS b FROM {table}', con
    )
    con.close()
    a, b = df.iloc[0]["a"], df.iloc[0]["b"]
    return (str(a)[:10], str(b)[:10]) if a else ("", "")


@st.cache_data(ttl=300, show_spinner="Đang truy vấn dữ liệu...")
def truy_van(table, date_col, hub_chon, loai_chon, tu, den, tim):
    dieu_kien, params = [], []
    if table == "quet_hang":
        if hub_chon:
            dieu_kien.append(f'"Hub" IN ({",".join(["%s"] * len(hub_chon))})')
            params += list(hub_chon)
        if loai_chon:
            dieu_kien.append(f'"Loại quét" IN ({",".join(["%s"] * len(loai_chon))})')
            params += list(loai_chon)
    if tu:
        dieu_kien.append(f'"{date_col}" >= %s')
        params.append(tu)
    if den:
        dieu_kien.append(f'"{date_col}" <= %s')
        params.append(den)
    if tim:
        ma_cot = {
            "quet_hang": '"Mã vận đơn"',
            "xep_hang":  '"Mã nhiệm vụ"',
        }[table]
        dieu_kien.append(f"{ma_cot} LIKE %s")
        params.append(f"%{tim}%")
    sql = f"SELECT * FROM {table}"
    if dieu_kien:
        sql += " WHERE " + " AND ".join(dieu_kien)
    sql += f' ORDER BY "{date_col}"'
    con = ket_noi()
    df = pd.read_sql(sql, con, params=params)
    con.close()

    if table == "quet_hang" and not df.empty:
        df["Mã chuẩn"] = _chuan_hoa_ma(df["Mã vận đơn"])
        danh_sach_mvd = list(set(df["Mã chuẩn"].tolist()) - {"\n"})

        if danh_sach_mvd:
            df_hist_list = []
            chunk_size = 50000
            con2 = ket_noi()
            for i in range(0, len(danh_sach_mvd), chunk_size):
                chunk = tuple(danh_sach_mvd[i : i + chunk_size])
                sql_history = """
                    SELECT "Mã vận đơn", "Thời gian quét", "Loại quét"
                    FROM quet_hang
                    WHERE "Mã vận đơn" IN %s
                      AND "Loại quét" IN ('Dỡ xuống xe', 'Xếp lên xe')
                """
                df_hist_list.append(pd.read_sql(sql_history, con2, params=(chunk,)))
            con2.close()

            df_hist = pd.concat(df_hist_list, ignore_index=True)
            df_hist["Mã chuẩn"] = _chuan_hoa_ma(df_hist["Mã vận đơn"])
            df_hist["Thời gian quét_dt"] = pd.to_datetime(df_hist["Thời gian quét"], errors="coerce")
            df_hist = df_hist.sort_values(by=["Mã chuẩn", "Thời gian quét_dt"])

            la_do_xuong = (df_hist["Loại quét"] == "Dỡ xuống xe").astype(int)
            df_hist["Đợt"] = la_do_xuong.groupby(df_hist["Mã chuẩn"]).cumsum()
            df_hist["Mã nối OB"] = df_hist["Mã chuẩn"] + "_Đợt" + df_hist["Đợt"].astype(str)

            la_xep_len = df_hist["Loại quét"] == "Xếp lên xe"
            df_hist["Trùng OB (sai thao tác)"] = False
            df_hist.loc[la_xep_len, "Trùng OB (sai thao tác)"] = df_hist[la_xep_len].duplicated(subset=["Mã nối OB"], keep="first")

            df_hist["Key_map"] = df_hist["Mã chuẩn"] + "_" + df_hist["Thời gian quét_dt"].dt.strftime('%Y%m%d%H%M%S')
            df["Thời gian quét_dt"] = pd.to_datetime(df["Thời gian quét"], errors="coerce")
            df["Key_map"] = df["Mã chuẩn"] + "_" + df["Thời gian quét_dt"].dt.strftime('%Y%m%d%H%M%S')

            df_hist_unique = df_hist.drop_duplicates(subset=["Key_map"], keep="last")
            map_ma_noi = df_hist_unique.set_index("Key_map")["Mã nối OB"].to_dict()
            map_trung = df_hist_unique.set_index("Key_map")["Trùng OB (sai thao tác)"].to_dict()

            df["Mã nối OB"] = df["Key_map"].map(map_ma_noi)
            df["Trùng OB (sai thao tác)"] = df["Key_map"].map(map_trung).fillna(False)
            df = df.drop(columns=["Mã chuẩn", "Thời gian quét_dt", "Key_map"])

    return df


@st.cache_data(ttl=300, show_spinner="Đang tính Ontime xếp hàng...")
@st.cache_data(ttl=300, show_spinner="Đang đọc COT từ Feishu...")
def lay_cutoff_feishu():
    """
    Đọc giờ COT trực tiếp từ Sheet COT trên Feishu.
    Không dùng bảng cutoff trên Supabase nữa.

    Trả về DataFrame gồm:
      - Bưu cục
      - OB
      - Hub (nếu sheet có cột Hub)
    """
    token = _feishu_token()
    if not token:
        raise RuntimeError("Không lấy được Feishu token để đọc Sheet COT.")

    url = (
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
        f"{FEISHU_COT_SPREADSHEET_TOKEN}/values/"
        f"{FEISHU_COT_SHEET_ID}!A1:Z10000"
        f"?valueRenderOption=ToString"
    )

    r = requests.get(
        url,
        headers=_feishu_headers(token),
        timeout=30
    )
    data = r.json()

    if data.get("code") != 0:
        raise RuntimeError(
            f"Không đọc được Sheet COT Feishu: {data.get('msg')}"
        )

    values = data.get("data", {}).get("valueRange", {}).get("values", [])
    if not values:
        raise RuntimeError("Sheet COT Feishu không có dữ liệu.")

    header = [str(x).strip() for x in values[0]]

    bc_idx = -1
    ob_idx = -1
    hub_idx = -1

    for i, val in enumerate(header):
        val_lower = val.lower()

        if (
            "bưu cục xuất" in val_lower
            or "bưu cục và hub" in val_lower
            or ("bưu cục" in val_lower and bc_idx == -1)
        ):
            bc_idx = i

        if val_lower == "ob":
            ob_idx = i

        if "hub" in val_lower:
            hub_idx = i

    # Fallback giống logic bộ cào JFS:
    if ob_idx == -1:
        for i, val in enumerate(header):
            if "ob" in val.lower():
                ob_idx = i
                break

    if bc_idx == -1 or ob_idx == -1:
        raise RuntimeError(
            f"Không tìm thấy cột Bưu cục/OB trong Sheet COT. "
            f"Header hiện tại: {header}"
        )

    rows = []
    for row in values[1:]:
        def val_at(idx):
            if idx < 0 or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        bưu_cuc = val_at(bc_idx)
        ob = val_at(ob_idx)
        hub_sheet = val_at(hub_idx)

        if not bưu_cuc or not ob:
            continue

        # Chuẩn hóa giờ COT về HH:MM:SS.
        if "." in ob and ob.replace(".", "").isdigit():
            try:
                val_float = float(ob)
                if 0 <= val_float <= 1:
                    total_seconds = round(val_float * 86400)
                    h, rem = divmod(total_seconds, 3600)
                    m, s = divmod(rem, 60)
                    ob = f"{int(h):02d}:{int(m):02d}:{int(s):02d}"
            except Exception:
                pass
        elif ":" in ob and len(ob.split(":")) == 2:
            ob += ":00"

        rows.append({
            "Bưu cục": bưu_cuc,
            "OB": ob,
            "Hub": hub_sheet,
        })

    if not rows:
        raise RuntimeError("Sheet COT không có dòng COT hợp lệ.")

    return pd.DataFrame(rows)


def tinh_ontime_xep_hang(hub, tu, den, loai_tuyen_chon=None):
    # COT được đọc trực tiếp từ Sheet COT Feishu,
    # không còn phụ thuộc bảng cutoff trên Supabase.
    try:
        cut = lay_cutoff_feishu()
    except Exception as e:
        return None, f"Không đọc được COT từ Feishu: {e}"

    bo_phan = TEN_BO_PHAN_XEP.get(hub, "")

    # Nếu Sheet COT có cột Hub thì ưu tiên đúng Hub đang truy vấn,
    # giữ nguyên nguyên tắc ưu tiên Hub của logic cũ trên Supabase.
    if "Hub" in cut.columns and cut["Hub"].astype(str).str.strip().ne("").any():
        cut["uu_tien"] = (
            cut["Hub"].astype(str).str.strip() == bo_phan
        ).astype(int)
        cut = cut.sort_values("uu_tien", ascending=False)

    cut = cut.drop_duplicates(subset="Bưu cục")[["Bưu cục", "OB"]]

    con = ket_noi()

    sql = 'SELECT * FROM xep_hang WHERE "Bộ phận xếp hàng" = %s'
    params = [bo_phan]
    if loai_tuyen_chon:
        sql += f' AND "Loại tuyến đường" IN ({",".join(["%s"] * len(loai_tuyen_chon))})'
        params += list(loai_tuyen_chon)
    if tu:
        sql += ' AND "Ngày vận hành" >= %s'
        params.append(tu)
    if den:
        sql += ' AND "Ngày vận hành" <= %s'
        params.append(den)
    df = pd.read_sql(sql, con, params=params)
    con.close()

    if df.empty:
        return None, f"Không có dữ liệu xếp hàng cho {bo_phan} trong khoảng ngày này."

    df["Loại tuyến đường"] = df["Loại tuyến đường"].map(
        lambda v: LOAI_TUYEN_NHAN.get(v, v)
    )

    df = df.merge(
        cut.rename(columns={"Bưu cục": "Bộ phận tiếp theo", "OB": "Cot"}),
        on="Bộ phận tiếp theo", how="left",
    )

    tg_bd  = pd.to_datetime(df["Thời gian bắt đầu xếp hàng"], errors="coerce")
    gio_bd  = tg_bd - tg_bd.dt.normalize()               
    gio_cot = pd.to_timedelta(df["Cot"], errors="coerce")     

    ngay_chuan = tg_bd.dt.normalize() + pd.to_timedelta(
        (gio_bd > gio_cot).astype(int), unit="D"
    )
    df["Thời gian giao hàng quy định"] = ngay_chuan + gio_cot

    tg_hoan_thanh = pd.to_datetime(df["Thời gian hoàn thành xếp hàng"], errors="coerce")
    tg_phat_xe    = pd.to_datetime(df["Thời gian phát xe"], errors="coerce")
    df["Thời gian giao hàng"] = tg_hoan_thanh.fillna(tg_phat_xe)

    df["Ontime"] = "Late"
    df.loc[
        df["Thời gian giao hàng"] < df["Thời gian giao hàng quy định"], "Ontime"
    ] = "Ontime"
    df.loc[
        df["Thời gian giao hàng"].notna() & df["Thời gian giao hàng quy định"].isna(),
        "Ontime",
    ] = "Thiếu cutoff"

    tn_mask = (
        df["Bộ phận tiếp theo"].astype(str).str.strip().str.upper().str.endswith("TN")
    )
    df.loc[tn_mask, "Ontime"] = "Không tính Ontime"

    return df, None



@st.cache_data(ttl=300, show_spinner="Đang lấy báo cáo Ontime...")
def truy_van_bao_cao_ontime(tu, den, hub_chon=None):
    """
    HCM/SHDC: giữ nguyên logic hiện tại.

    BN:
      - Mẫu số: DISTINCT Mã vận đơn của quet_hang / BN / Dỡ xuống xe.
      - Tử số: các dòng bao_cao có Trạng thái xử lý = Ontime trong ngày,
        nhưng đối chiếu theo Mã vận đơn + Ngày vận hành với đúng tập BN
        ở mẫu số. Không lọc trùng lại bao_cao.
    """
    chi_bn = (
        len(hub_chon) == 1
        and str(hub_chon[0]).strip() == "BN"
    )

    if chi_bn:
        # 1) Lấy đúng tập mã vận đơn BN làm mẫu số.
        sql_bn = """
            SELECT "Mã vận đơn", "Ngày vận hành"
            FROM quet_hang
            WHERE "Hub" = %s
              AND "Loại quét" = %s
              AND "Ngày vận hành" >= %s
              AND "Ngày vận hành" <= %s
        """

        con = ket_noi()
        try:
            df_bn = pd.read_sql(
                sql_bn,
                con,
                params=("BN", "Dỡ xuống xe", tu, den),
            )
        finally:
            con.close()

        if df_bn.empty:
            bn_keys = set()
        else:
            df_bn["Mã chuẩn"] = _chuan_hoa_ma(df_bn["Mã vận đơn"])
            df_bn["Ngày vận hành"] = df_bn["Ngày vận hành"].astype(str)
            df_bn = df_bn[df_bn["Mã chuẩn"] != ""].copy()
            bn_keys = set(
                zip(df_bn["Mã chuẩn"], df_bn["Ngày vận hành"])
            )

        tong_mau_so = len(
            {mvd for mvd, _ in bn_keys}
        )

        # 2) Lấy TOÀN BỘ bao_cao theo ngày + trạng thái Ontime.
        # Không filter Hub ở đây để không làm mất các BN đang có nhãn Hub
        # khác trong bao_cao.
        sql_ontime = """
            SELECT *
            FROM bao_cao
            WHERE "Trạng thái xử lý" = %s
              AND "Ngày vận hành" >= %s
              AND "Ngày vận hành" <= %s
            ORDER BY "Ngày vận hành", "Thời gian quét"
        """

        con = ket_noi()
        try:
            df_ontime = pd.read_sql(
                sql_ontime,
                con,
                params=("Ontime", tu, den),
            )
        finally:
            con.close()

        if df_ontime.empty or not bn_keys:
            df_ontime = df_ontime.iloc[0:0].copy()
        else:
            df_ontime["Mã chuẩn"] = _chuan_hoa_ma(
                df_ontime["Mã vận đơn"]
            )
            df_ontime["Ngày vận hành"] = (
                df_ontime["Ngày vận hành"].astype(str)
            )

            # Chỉ đối chiếu tập mã/ngày của BN; không dedup lại Ontime.
            df_ontime = df_ontime.loc[
                list(
                    zip(
                        df_ontime["Mã chuẩn"],
                        df_ontime["Ngày vận hành"],
                    )
                )
            ].copy()

            keep_mask = [
                key in bn_keys
                for key in zip(
                    df_ontime["Mã chuẩn"],
                    df_ontime["Ngày vận hành"],
                )
            ]
            df_ontime = df_ontime.loc[keep_mask].copy()
            df_ontime = df_ontime.drop(columns=["Mã chuẩn"])

        df_all = df_ontime.copy()
        df_ontime.attrs["tong_mau_so"] = tong_mau_so
        return df_all, df_ontime

    # ---------------- HCM / SH DC: giữ nguyên logic hiện tại ----------------
    dieu_kien = []
    params = []

    if tu:
        dieu_kien.append('"Ngày vận hành" >= %s')
        params.append(tu)

    if den:
        dieu_kien.append('"Ngày vận hành" <= %s')
        params.append(den)

    if hub_chon:
        hub_db = [
            HUB_BAO_CAO.get(str(h).strip(), str(h).strip())
            for h in hub_chon
        ]
        placeholders = ",".join(["%s"] * len(hub_db))
        dieu_kien.append(f'"Hub" IN ({placeholders})')
        params.extend(hub_db)

    sql = 'SELECT * FROM bao_cao'
    if dieu_kien:
        sql += ' WHERE ' + ' AND '.join(dieu_kien)
    sql += ' ORDER BY "Ngày vận hành", "Thời gian quét"'

    con = ket_noi()
    try:
        df_all = pd.read_sql(sql, con, params=params)
    finally:
        con.close()

    if df_all.empty:
        df_all.attrs["tong_mau_so"] = 0
        return df_all, df_all.copy()

    if "Kết hợp" in df_all.columns:
        df_all = df_all.drop_duplicates(
            subset=["Kết hợp"],
            keep="last",
        ).copy()

    trang_thai = df_all["Trạng thái xử lý"].astype(str).str.strip()
    df_ontime = df_all.loc[trang_thai.eq("Ontime")].copy()

    df_ontime.attrs["tong_mau_so"] = len(df_all)
    return df_all, df_ontime



def _feishu_ghi_bao_cao_ontime(token, sheet_name, tong_mau_so,
                                tong_ontime, ty_le_ontime, df_ontime):
    """
    Ghi một sheet Feishu:
      - KPI ở đầu sheet
      - toàn bộ chi tiết đơn Ontime ở phía dưới.
    """
    sheet_id = _feishu_lay_hoac_tao_sheet(token, sheet_name)
    _feishu_don_dep_sheet(token, sheet_id)

    kpi_rows = [
        ["Chỉ tiêu", "Giá trị", "Tỷ lệ"],
        ["Tổng đơn Dỡ xuống xe sau lọc trùng", int(tong_mau_so), "100.00%"],
        ["Tổng đơn Ontime", int(tong_ontime), f"{ty_le_ontime:.2f}%"],
        [],
        ["CHI TIẾT TẤT CẢ ĐƠN ONTIME"],
    ]

    detail = _df_ve_gia_tri(df_ontime)
    values = kpi_rows + detail

    if not values:
        raise RuntimeError("Không có dữ liệu để ghi lên Feishu.")

    max_cols = max(len(r) for r in values)
    values = [
        list(r) + [""] * (max_cols - len(r))
        for r in values
    ]

    from openpyxl.utils import get_column_letter

    cot_kt = get_column_letter(max_cols)
    url = (
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
        f"{FEISHU_SPREADSHEET_TOKEN}/values"
    )

    CHUNK = 500
    for i in range(0, len(values), CHUNK):
        phan = values[i:i + CHUNK]
        dong_bd = i + 1
        dong_kt = i + len(phan)
        rng = f"{sheet_id}!A{dong_bd}:{cot_kt}{dong_kt}"

        body = {
            "valueRange": {
                "range": rng,
                "values": phan,
            }
        }

        r = requests.put(
            url,
            headers=_feishu_headers(token),
            json=body,
            timeout=60,
        )
        d = r.json()

        if d.get("code") != 0:
            raise RuntimeError(
                f"Lỗi ghi báo cáo Ontime lên Feishu: {d.get('msg')}"
            )

    return sheet_id


def day_len_bao_cao_ontime(sheet_name, df_all, df_ontime):
    if df_all is None or df_all.empty:
        raise RuntimeError("Không có dữ liệu bao_cao trong điều kiện đã chọn.")

    token = _feishu_token()

    tong_mau_so = len(df_all)
    tong_ontime = len(df_ontime)
    ty_le_ontime = (
        tong_ontime / tong_mau_so * 100
        if tong_mau_so else 0
    )

    return _feishu_ghi_bao_cao_ontime(
        token,
        sheet_name,
        tong_mau_so,
        tong_ontime,
        ty_le_ontime,
        df_ontime,
    )


@st.cache_data(ttl=300, show_spinner="Đang tính Ontime 1AM...")
def truy_van_ontime_1am(tu, den, hub_chon=None):
    """
    Ontime 1AM.

    Mẫu số:
      - Bảng quet_hang
      - Loại quét = Dỡ xuống xe
      - Scan trước 01:00 của ngày kế tiếp so với Ngày vận hành
      - Đếm đơn duy nhất theo Mã vận đơn + Ngày vận hành

    Tử số:
      - Bảng bao_cao
      - Trạng thái xử lý = Ontime
      - IB trước 1h sáng = IB trước 01:00
    """
    where_parts = [
        '"Loại quét" = %s',
        '"Ngày vận hành" >= %s',
        '"Ngày vận hành" <= %s',
    ]
    params = ["Dỡ xuống xe", tu, den]

    if hub_chon:
        placeholders = ",".join(["%s"] * len(hub_chon))
        where_parts.append(f'"Hub" IN ({placeholders})')
        params.extend(list(hub_chon))

    sql_ib = (
        'SELECT * FROM quet_hang WHERE '
        + " AND ".join(where_parts)
        + ' ORDER BY "Ngày vận hành", "Thời gian quét"'
    )

    con = ket_noi()
    try:
        df_ib = pd.read_sql(sql_ib, con, params=params)
    finally:
        con.close()

    if not df_ib.empty:
        df_ib["Mã chuẩn"] = _chuan_hoa_ma(df_ib["Mã vận đơn"])
        df_ib["Thời gian quét_dt"] = pd.to_datetime(
            df_ib["Thời gian quét"], errors="coerce"
        )
        df_ib["Ngày vận hành_dt"] = pd.to_datetime(
            df_ib["Ngày vận hành"], errors="coerce"
        )

        # 01:00 của ngày kế tiếp theo Ngày vận hành.
        df_ib["gioi_han_1am"] = (
            df_ib["Ngày vận hành_dt"]
            + pd.Timedelta(days=1, hours=1)
        )

        df_ib = df_ib[
            df_ib["Thời gian quét_dt"] < df_ib["gioi_han_1am"]
        ].copy()

        df_ib = (
            df_ib[df_ib["Mã chuẩn"] != ""]
            .drop_duplicates(
                subset=["Mã chuẩn", "Ngày vận hành"],
                keep="last",
            )
        )

    tong_ib_1am = len(df_ib)

    where_bc = [
        '"Trạng thái xử lý" = %s',
        '"IB trước 1h sáng" = %s',
        '"Ngày vận hành" >= %s',
        '"Ngày vận hành" <= %s',
    ]
    params_bc = ["Ontime", "IB trước 01:00", tu, den]

    if hub_chon:
        # Giao diện dùng HCM / BN / SH DC, nhưng bao_cao lưu HCM HUB / BN HUB / SH DC.
        hub_db = [HUB_BAO_CAO.get(str(h).strip(), str(h).strip()) for h in hub_chon]
        placeholders = ",".join(["%s"] * len(hub_db))
        where_bc.append(f'"Hub" IN ({placeholders})')
        params_bc.extend(hub_db)

    sql_bc = (
        'SELECT * FROM bao_cao WHERE '
        + " AND ".join(where_bc)
        + ' ORDER BY "Ngày vận hành", "Thời gian quét"'
    )

    con = ket_noi()
    try:
        df_ontime_1am = pd.read_sql(sql_bc, con, params=params_bc)
    finally:
        con.close()

    if not df_ontime_1am.empty and "Kết hợp" in df_ontime_1am.columns:
        df_ontime_1am = df_ontime_1am.drop_duplicates(
            subset=["Kết hợp"],
            keep="last",
        ).copy()

    tong_ontime_1am = len(df_ontime_1am)
    ty_le_ontime_1am = (
        tong_ontime_1am / tong_ib_1am * 100
        if tong_ib_1am else 0
    )

    return tong_ib_1am, tong_ontime_1am, ty_le_ontime_1am, df_ontime_1am


@st.cache_data(ttl=300, show_spinner="Đang tạo file Excel...")
def xuat_excel(df_dict):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as wr:
        for ten, d in df_dict.items():
            d.to_excel(wr, sheet_name=ten, index=False)
    return buf.getvalue()


def _feishu_token():
    if not FEISHU_APP_ID or not FEISHU_APP_SECRET:
        raise RuntimeError("Chưa khai báo FEISHU_APP_ID / FEISHU_APP_SECRET.")
    r = requests.post(
        "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
        json={"app_id": FEISHU_APP_ID, "app_secret": FEISHU_APP_SECRET},
        timeout=15,
    )
    data = r.json()
    if data.get("code") != 0:
        raise RuntimeError(f"Feishu từ chối cấp token: {data.get('msg')}")
    return data["tenant_access_token"]


def _feishu_headers(token):
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json; charset=utf-8",
    }


def _feishu_lay_hoac_tao_sheet(token, sheet_name):
    if not FEISHU_SPREADSHEET_TOKEN:
        raise RuntimeError("Chưa khai báo FEISHU_SPREADSHEET_TOKEN.")
    url = f"https://open.feishu.cn/open-apis/sheets/v3/spreadsheets/{FEISHU_SPREADSHEET_TOKEN}/sheets/query"
    r = requests.get(url, headers=_feishu_headers(token), timeout=15)
    data = r.json()
    if data.get("code") != 0:
        raise RuntimeError(f"Không lấy được danh sách sheet: {data.get('msg')}")
    for s in data["data"]["sheets"]:
        if s["title"] == sheet_name:
            return s["sheet_id"]
    url_add = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{FEISHU_SPREADSHEET_TOKEN}/sheets_batch_update"
    body = {"requests": [{"addSheet": {"properties": {"title": sheet_name, "index": 0}}}]}
    r2 = requests.post(url_add, headers=_feishu_headers(token), json=body, timeout=15)
    d2 = r2.json()
    if d2.get("code") != 0:
        raise RuntimeError(f"Không tạo được sheet '{sheet_name}': {d2.get('msg')}")
    return d2["data"]["replies"][0]["addSheet"]["properties"]["sheetId"]


def _feishu_don_dep_sheet(token, sheet_id):
    url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{FEISHU_SPREADSHEET_TOKEN}/values_batch_clear"
    body = {"ranges": [f"{sheet_id}!A1:BZ200000"]}
    requests.post(url, headers=_feishu_headers(token), json=body, timeout=30)


def _df_ve_gia_tri(df):
    d = df.copy()
    for c in d.columns:
        if pd.api.types.is_datetime64_any_dtype(d[c]):
            d[c] = d[c].astype(str)
    d = d.astype(object).where(pd.notnull(d), "")
    d = d.astype(str)
    return [list(d.columns)] + d.values.tolist()


def day_len_feishu(sheet_name, df):
    if df is None or df.empty:
        raise RuntimeError("Không có dữ liệu để đẩy.")
    token = _feishu_token()
    sheet_id = _feishu_lay_hoac_tao_sheet(token, sheet_name)
    _feishu_don_dep_sheet(token, sheet_id)
    gia_tri = _df_ve_gia_tri(df)
    so_cot = len(gia_tri[0])
    tong_dong = len(gia_tri)
    CHUNK = 2000
    url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{FEISHU_SPREADSHEET_TOKEN}/values"
    cot_kt = get_column_letter(so_cot)
    for i in range(0, tong_dong, CHUNK):
        phan = gia_tri[i : i + CHUNK]
        dong_bd = i + 1
        dong_kt = i + len(phan)
        rng = f"{sheet_id}!A{dong_bd}:{cot_kt}{dong_kt}"
        body = {"valueRange": {"range": rng, "values": phan}}
        r = requests.put(url, headers=_feishu_headers(token), json=body, timeout=60)
        d = r.json()
        if d.get("code") != 0:
            raise RuntimeError(f"Lỗi ghi dữ liệu: {d.get('msg')}")
    return tong_dong - 1


def divider_label(text):
    st.markdown(
        f'<div style="font-size:12px;font-weight:700;color:#6c757d;'
        f'text-transform:uppercase;letter-spacing:.06em;'
        f'border-left:3px solid #1a3a6b;padding-left:8px;margin:18px 0 10px;">'
        f'{text}</div>',
        unsafe_allow_html=True,
    )


# ── Tabs đúng nguyên bản: Dữ liệu thô & Ontime Xếp hàng ────────────────────────
tab_du_lieu, tab_ontime_xh, tab_bao_cao_ontime = st.tabs([
    "  Dữ liệu thô  ",
    "  Ontime Xếp hàng (Xe)  ",
    "  Báo cáo Ontime  ",
])


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 — DỮ LIỆU THÔ (Quét hàng)
# ═══════════════════════════════════════════════════════════════════════════════
with tab_du_lieu:
    with st.form("bo_loc"):
        c1, c2, c3 = st.columns(3)
        ten_bang = c1.selectbox("Bảng dữ liệu", list(BANG.keys()))
        table, date_col = BANG[ten_bang]
        hub_chon = c2.multiselect("Hub", lay_hub(), placeholder="Tất cả")
        loai_chon = c3.multiselect(
            "Loại quét", ["Xếp lên xe", "Dỡ xuống xe"], placeholder="Tất cả"
        )
        c4, c5, c6 = st.columns(3)
        ngay_min, ngay_max = khoang_ngay(table, date_col)
        tu = c4.date_input("Từ ngày", value=_ngay(ngay_min) or datetime.date.today(), format="YYYY-MM-DD")
        den = c5.date_input("Đến ngày", value=_ngay(ngay_max) or datetime.date.today(), format="YYYY-MM-DD")
        tim = c6.text_input("Tìm mã vận đơn").strip()
        truy_van_clicked = st.form_submit_button(
            "Truy vấn", type="primary", use_container_width=True
        )

    if truy_van_clicked:
        st.session_state["kq"] = truy_van(
            table, date_col, tuple(hub_chon), tuple(loai_chon),
            str(tu) if tu else "", str(den) if den else "", tim,
        )
        st.session_state["bang_kq"] = table

    if "kq" in st.session_state:
        df = st.session_state["kq"]
        table = st.session_state["bang_kq"]

        divider_label("Kết quả")

        if table == "quet_hang" and not df.empty:
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Số dòng", f"{len(df):,}")
            
            if "Mã nối OB" in df.columns:
                la_ob_dem = df["Loại quét"] == "Xếp lên xe"
                khoa_dem = df["Mã vận đơn"].astype(str).where(~la_ob_dem, df["Mã nối OB"])
                so_van_don = khoa_dem.nunique()
            else:
                so_van_don = df["Mã vận đơn"].nunique()
            m2.metric("Số vận đơn", f"{so_van_don:,}")

            mask_ob_w = df["Loại quét"] == "Xếp lên xe"
            mask_ib_w = df["Loại quét"] == "Dỡ xuống xe"

            if "Mã nối OB" in df.columns:
                trong_luong_ob = df[mask_ob_w].drop_duplicates(subset="Mã nối OB")["Trọng lượng"].sum()
            else:
                trong_luong_ob = df.loc[mask_ob_w, "Trọng lượng"].sum()

            trong_luong_ib = df[mask_ib_w].drop_duplicates(subset="Mã vận đơn")["Trọng lượng"].sum()
            trong_luong_tong = trong_luong_ob + trong_luong_ib

            m3.metric("Trọng lượng (kg)", f"{trong_luong_tong:,.0f}")
            so_trung_ob = df["Trùng OB (sai thao tác)"].sum() if "Trùng OB (sai thao tác)" in df.columns else 0
            m4.metric("Trùng OB (sai thao tác)", f"{so_trung_ob:,}")
        else:
            st.metric("Số dòng", f"{len(df):,}")

        st.dataframe(df.head(SO_DONG_HIEN_THI), use_container_width=True, height=440)

        if table == "quet_hang" and not df.empty:
            if "Mã nối OB" in df.columns:
                la_ob_xuat = df["Loại quét"] == "Xếp lên xe"
                khoa_xuat = df["Mã vận đơn"].astype(str).where(~la_ob_xuat, df["Mã nối OB"])
                df_xuat = df.loc[~khoa_xuat.duplicated(keep="first")]
            else:
                df_xuat = df.drop_duplicates(subset="Mã vận đơn")
            nhan_tai = f"Tải Excel ({len(df_xuat):,} số vận đơn)"
        else:
            df_xuat = df
            nhan_tai = f"Tải Excel ({len(df_xuat):,} dòng)"

        cot_tai, cot_day = st.columns(2)
        with cot_tai:
            if 0 < len(df_xuat) <= 1_000_000:
                st.download_button(
                    nhan_tai,
                    xuat_excel({"Du lieu": df_xuat}),
                    file_name=f"{table}.xlsx",
                    use_container_width=True,
                    type="primary",
                )


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 — ONTIME XẾP HÀNG (Quản lý tiến độ xếp xe)
# ═══════════════════════════════════════════════════════════════════════════════
with tab_ontime_xh:
    with st.form("ontime_xh_form"):
        c1, c2, c3 = st.columns(3)
        hub_xh = c1.selectbox("Hub", lay_hub(), key="hub_xh")
        nmin_xh, nmax_xh = khoang_ngay("xep_hang", "Ngày vận hành")
        tu_xh = c2.date_input("Từ ngày vận hành", value=_ngay(nmin_xh) or datetime.date.today(), format="YYYY-MM-DD")
        den_xh = c3.date_input("Đến ngày vận hành", value=_ngay(nmax_xh) or datetime.date.today(), format="YYYY-MM-DD")
        loai_tuyen_xh = st.multiselect(
            "Loại tuyến đường",
            list(LOAI_TUYEN_NHAN.values()),
            placeholder="Tất cả (Tuyến chính + Tuyến phụ)",
        )
        truy_van_xh_clicked = st.form_submit_button(
            "Truy vấn", type="primary", use_container_width=True
        )

    if truy_van_xh_clicked:
        loai_tuyen_thuc = tuple(
            NHAN_TOI_LOAI_TUYEN.get(nhan, nhan) for nhan in loai_tuyen_xh
        )
        kq_xh, loi_xh = tinh_ontime_xep_hang(
            hub_xh,
            str(tu_xh) if tu_xh else "",
            str(den_xh) if den_xh else "",
            loai_tuyen_thuc,
        )
        if loi_xh:
            st.error(loi_xh)
        else:
            st.session_state["ontime_xh"] = kq_xh
            st.session_state["ontime_xh_hub"] = hub_xh

    if "ontime_xh" in st.session_state:
        kq_xh = st.session_state["ontime_xh"]
        hub_xh_hien_tai = st.session_state["ontime_xh_hub"]

        kq_tinh = kq_xh[kq_xh["Ontime"] != "Không tính Ontime"]

        lh_mask = kq_tinh["Loại tuyến đường"].str.contains("Linehaul", na=False)
        sl_mask = kq_tinh["Loại tuyến đường"].str.contains("Shuttle", na=False)

        tong_lh = lh_mask.sum()
        tong_shuttle = sl_mask.sum()
        tre_lh = (lh_mask & (kq_tinh["Ontime"] == "Late")).sum()
        tre_shuttle = (sl_mask & (kq_tinh["Ontime"] == "Late")).sum()

        so_ontime_xh = (kq_tinh["Ontime"] == "Ontime").sum()
        ty_le_xh = so_ontime_xh / len(kq_tinh) * 100 if len(kq_tinh) else 0

        divider_label("Tổng quan")
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Tổng LH", f"{tong_lh:,}")
        c2.metric("Tổng Shuttle", f"{tong_shuttle:,}")
        c3.metric("Trễ LH", f"{tre_lh:,}")
        c4.metric("Trễ Shuttle", f"{tre_shuttle:,}")
        c5.metric("Tỷ lệ Ontime", f"{ty_le_xh:.2f}%")

        tong_hop_xh = (
            kq_xh.groupby(["Ngày vận hành", "Loại tuyến đường", "Ontime"]).size()
            .unstack(fill_value=0).reset_index()
        )

        divider_label("Tổng hợp theo ngày vận hành + loại tuyến")
        st.dataframe(tong_hop_xh, use_container_width=True, height=280)

        divider_label(f"Chi tiết Ontime xếp hàng ({min(len(kq_xh), SO_DONG_HIEN_THI):,} / {len(kq_xh):,} dòng)")
        cot_hien_thi = [
            "Mã nhiệm vụ", "Loại tuyến đường", "Bộ phận xếp hàng", "Bộ phận tiếp theo",
            "Thời gian bắt đầu xếp hàng", "Cot", "Thời gian giao hàng quy định",
            "Thời gian hoàn thành xếp hàng", "Thời gian phát xe",
            "Thời gian giao hàng", "Ontime", "Ngày vận hành",
        ]
        cot_hien_thi = [c for c in cot_hien_thi if c in kq_xh.columns]
        st.dataframe(
            kq_xh[cot_hien_thi].head(SO_DONG_HIEN_THI),
            use_container_width=True, height=320,
        )

        st.download_button(
            f"Tải Excel — Ontime xếp hàng ({len(kq_xh):,} nhiệm vụ)",
            xuat_excel({"Ontime xep hang": kq_xh, "Tong hop": tong_hop_xh}),
            file_name=f"ontime_xep_hang_{hub_xh_hien_tai}.xlsx",
            use_container_width=True,
            type="primary",
        )


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3 — BÁO CÁO ONTIME (DỠ XUỐNG XE)
# ═══════════════════════════════════════════════════════════════════════════════
with tab_bao_cao_ontime:
    with st.form("bao_cao_ontime_form"):
        c1, c2, c3 = st.columns(3)

        nmin_bc, nmax_bc = khoang_ngay("bao_cao", "Ngày vận hành")

        tu_bc = c1.date_input(
            "Từ ngày vận hành",
            value=_ngay(nmin_bc) or datetime.date.today(),
            format="YYYY-MM-DD",
        )
        den_bc = c2.date_input(
            "Đến ngày vận hành",
            value=_ngay(nmax_bc) or datetime.date.today(),
            format="YYYY-MM-DD",
        )
        hub_bc = c3.multiselect(
            "Hub",
            lay_hub(),
            placeholder="Tất cả",
            key="hub_bc_ontime",
        )

        truy_van_bc_clicked = st.form_submit_button(
            "Truy vấn báo cáo Ontime",
            type="primary",
            use_container_width=True,
        )

    if truy_van_bc_clicked:
        try:
            df_bc_all, df_bc_ontime = truy_van_bao_cao_ontime(
                str(tu_bc) if tu_bc else "",
                str(den_bc) if den_bc else "",
                tuple(hub_bc),
            )

            st.session_state["bc_ontime_all"] = df_bc_all
            st.session_state["bc_ontime_detail"] = df_bc_ontime
            st.session_state["bc_ontime_mau_so"] = df_bc_ontime.attrs.get(
                "tong_mau_so",
                len(df_bc_all),
            )
            st.session_state["bc_ontime_tu"] = str(tu_bc) if tu_bc else ""
            st.session_state["bc_ontime_den"] = str(den_bc) if den_bc else ""
            st.session_state["bc_ontime_hub"] = tuple(hub_bc)

        except Exception as e:
            st.error(f"Lỗi truy vấn báo cáo Ontime: {e}")

    if "bc_ontime_all" in st.session_state:
        df_bc_all = st.session_state["bc_ontime_all"]
        df_bc_ontime = st.session_state["bc_ontime_detail"]

        tong_mau_so = st.session_state.get(
            "bc_ontime_mau_so",
            len(df_bc_all),
        )
        tong_ontime = len(df_bc_ontime)
        ty_le_ontime = (
            tong_ontime / tong_mau_so * 100
            if tong_mau_so else 0
        )

        divider_label("Tổng quan")

        c1, c2, c3 = st.columns(3)

        c1.metric(
            "Tổng đơn Dỡ xuống xe sau lọc trùng",
            f"{tong_mau_so:,}",
        )
        c2.metric(
            "Tổng đơn Ontime",
            f"{tong_ontime:,}",
        )
        c3.metric(
            "Tỷ lệ Ontime",
            f"{ty_le_ontime:.2f}%",
        )

        st.caption(
            f"Công thức: {tong_ontime:,} / {tong_mau_so:,} × 100 "
            f"= {ty_le_ontime:.2f}%"
        )

        # ── KPI: Ontime 1AM ──
        try:
            (
                tong_ib_1am,
                tong_ontime_1am,
                ty_le_ontime_1am,
                df_ontime_1am,
            ) = truy_van_ontime_1am(
                st.session_state.get("bc_ontime_tu", ""),
                st.session_state.get("bc_ontime_den", ""),
                st.session_state.get("bc_ontime_hub", ()),
            )

            divider_label("Ontime 1AM")

            c1, c2, c3 = st.columns(3)
            c1.metric("Tổng đơn IB trước 1AM", f"{tong_ib_1am:,}")
            c2.metric("Tổng đơn Ontime 1AM", f"{tong_ontime_1am:,}")
            c3.metric("Tỷ lệ Ontime 1AM", f"{ty_le_ontime_1am:.2f}%")

            st.caption(
                f"Công thức: {tong_ontime_1am:,} / {tong_ib_1am:,} × 100 "
                f"= {ty_le_ontime_1am:.2f}%"
            )
        except Exception as e:
            st.error(f"Lỗi tính Ontime 1AM: {e}")

        divider_label(
            f"Tất cả đơn Ontime "
            f"({min(len(df_bc_ontime), SO_DONG_HIEN_THI):,} / "
            f"{len(df_bc_ontime):,} dòng)"
        )

        st.dataframe(
            df_bc_ontime.head(SO_DONG_HIEN_THI),
            use_container_width=True,
            height=440,
        )

        st.download_button(
            f"Tải Excel — Tất cả đơn Ontime ({len(df_bc_ontime):,})",
            xuat_excel({"Ontime": df_bc_ontime}),
            file_name="bao_cao_ontime.xlsx",
            use_container_width=True,
            type="primary",
        )
